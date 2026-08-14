#!/usr/bin/env python3
"""
Build the LeadForge Excel workbook from leads/queue.jsonl.
Business = grouped rows (MAIN contact shown), with a Contacts sheet listing every contact
(person) per business for vertical outreach. Color-coded by stage.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LEADS = os.path.join(ROOT, "leads", "queue.jsonl")
OUT = os.path.join(ROOT, "leads", "LeadForge_Leads.xlsx")

STAGE_FILL = {
    "s1": "FFF2CC", "s2": "FCE5CD", "s3": "D9EAD3", "s4": "B6D7A8",
}
STAGE_LABEL = {"s1": "S1 Identified", "s2": "S2 Contact", "s3": "S3 Qualified", "s4": "S4 Hot"}


def load():
    out = []
    with open(LEADS) as f:
        for line in f:
            line = line.strip()
            if line:
                try: out.append(json.loads(line))
                except: pass
    return out


def main():
    leads = load()
    # group businesses
    biz = {}
    for l in leads:
        bid = l.get("business_id") or l.get("business", "")
        biz.setdefault(bid, []).append(l)

    wb = Workbook()
    # Businesses sheet
    ws = wb.active
    ws.title = "Businesses"
    headers = ["Business", "Vertical", "Town", "State", "Main Contact", "Title",
               "Business Phone", "Main Email", "Personal Cell", "Pain", "Stage",
               "Contacts", "Score", "Source"]
    ws.append(headers)
    for bid, rows in biz.items():
        main = next((r for r in rows if r.get("role") == "main"), rows[0])
        ws.append([
            main.get("business", ""), main.get("vertical", ""), main.get("town", ""),
            main.get("state", ""), main.get("first", "") or "—", main.get("title", ""),
            main.get("phone", "") or "—", main.get("email", "") or "—",
            main.get("cell", "") or "—", main.get("pain", "") or "—",
            STAGE_LABEL.get(main.get("stage", "s1"), main.get("stage", "")),
            len(rows), main.get("score", ""), main.get("source", ""),
        ])
        fill = STAGE_FILL.get(main.get("stage", "s1"), "FFFFFF")
        ws.cell(row=ws.max_row, column=11).fill = PatternFill("solid", fgColor=fill)

    # Contacts sheet (one row per person)
    ws2 = wb.create_sheet("Contacts")
    ch = ["Business", "Role", "Name", "Title", "Phone", "Cell", "Email",
          "Confidence", "Sources", "Stage", "Notes"]
    ws2.append(ch)
    for bid, rows in biz.items():
        for r in rows:
            ws2.append([
                r.get("business", ""), r.get("role", ""), r.get("first", "") or "—",
                r.get("title", ""), r.get("phone", "") or "—", r.get("cell", "") or "—",
                r.get("email", "") or "—", r.get("confidence", "") or "—",
                "; ".join(r.get("sources", [])), STAGE_LABEL.get(r.get("stage","s1"), ""),
                r.get("notes", ""),
            ])

    # Summary sheet
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Value"])
    ws3.append(["Total businesses", len(biz)])
    ws3.append(["Total contacts", len(leads)])
    for st in ("s1", "s2", "s3", "s4"):
        ws3.append([STAGE_LABEL[st], sum(1 for l in leads if l.get("stage") == st)])
    ws3.append(["With verified name", sum(1 for l in leads if l.get("first"))])
    ws3.append(["With personal cell", sum(1 for l in leads if l.get("cell"))])
    ws3.append(["With email", sum(1 for l in leads if l.get("email"))])

    # styling
    thin = Side(style="thin", color="DDDDDD")
    for sheet in (ws, ws2, ws3):
        for c in sheet[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F6FEB")
            c.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows():
            for c in row:
                c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # widths
    for sheet, w in ((ws, 22), (ws2, 20)):
        for col in sheet.columns:
            sheet.column_dimensions[col[0].column_letter].width = w

    wb.save(OUT)
    print(f"Wrote {OUT}: {len(biz)} businesses, {len(leads)} contacts.")


if __name__ == "__main__":
    main()
